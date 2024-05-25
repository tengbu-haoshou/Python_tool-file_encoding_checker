#!/usr/bin/env python3

#
# file_encoding_checker.py
#
# Date    : 2024-04-30
# Auther  : Hirotoshi FUJIBE
# History :
#
# Copyright (c) 2024 Hirotoshi FUJIBE
#

"""
Usage:

    Python.exe file_encoding_checker.py

Options:

    -h
    --help
        Print this message and exit.
"""

# Import Libraries
import os
import sys
import getopt
import shutil
import datetime
import openpyxl
from typing import Union
from chardet.universaldetector import UniversalDetector
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.styles.borders import Border, Side

# Input, Output
IN_DIR = '.\\input'
OUT_DIR = '.\\output'
# IN_SRC_ROOT = 'D:\\Developments\\PyCharmProjects\\tool-file_encoding_checker\\input'  # noqa
IN_SRC_ROOT = '.\\input'
IN_SRC_RELATIVE = '\\src'
IN_EXCEL = IN_DIR + '\\file_encoding_checker_list_template.xlsx'
OUT_EXCEL = OUT_DIR + '\\file_encoding_checker_list.xlsx'
OUT_SHEET = 'File Encoding Checker List'
IGNORE_EXTENDS = ['.dat', '.ini']
OUT_DEBUG = OUT_DIR + '\\debug.txt'

# Excel Cell Position (1 Origin)
CELL_ROW_OFFSET = 4
CELL_COL_NO = 2
CELL_COL_PATH = 3
CELL_COL_FILE = 4
CELL_COL_EXT = 5
CELL_COL_ENCODING = 6
CELL_COL_ATTERS = 7

# Output Excel Cell Format
ALIGN_LEFT = Alignment(horizontal='left', vertical='top', wrap_text=True)
ALIGN_LEFT_NO_WRAP = Alignment(horizontal='left', vertical='top', wrap_text=False)
ALIGN_CENTER = Alignment(horizontal='center', vertical='top', wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal='right', vertical='top', wrap_text=True)
FONT_MEIRYO = Font(name='Meiryo UI', size=10, color='000000')
FONT_MEIRYO_GRAY = Font(name='Meiryo UI', size=10, color='C0C0C0')
FONT_MEIRYO_BOLD = Font(name='Meiryo UI', size=10, color='000000', bold=True)
FILL_BRIGHT_GRAY = PatternFill(patternType='solid', fgColor='EBECF0')
NUMBER_FORMAT = '#,##0_ '
BORDER_ALL = Border(
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000'),
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'))


# Write Excel
class WriteExcel:

    def __init__(self, in_excel: str, out_excel: str, out_sheet: str) -> None:
        shutil.copy(in_excel, out_excel)
        self._wb = openpyxl.load_workbook(out_excel)
        self._sheet = self._wb[out_sheet]
        self._row_offset = CELL_ROW_OFFSET
        self._row = 0
        self._out_excel = out_excel
        return

    def next_row(self) -> None:
        self._row += 1
        return

    def get_count(self) -> int:
        return self._row + 1

    def write_cell(self, i_col: int, i_value: Union[int, str],
                   i_align: int = None, i_font: Font = None, i_format: str = None) -> None:
        self._sheet.cell(row=self._row_offset + self._row, column=i_col).border = BORDER_ALL
        if i_value is not None:
            self._sheet.cell(row=self._row_offset + self._row, column=i_col).value = i_value
        if i_align is not None:
            self._sheet.cell(row=self._row_offset + self._row, column=i_col).alignment = i_align
        if i_font is not None:
            self._sheet.cell(row=self._row_offset + self._row, column=i_col).font = i_font
        else:
            self._sheet.cell(row=self._row_offset + self._row, column=i_col).font = FONT_MEIRYO
        if i_format is not None:
            self._sheet.cell(row=self._row_offset + self._row, column=i_col).number_format = i_format
        return

    def close(self) -> None:
        self._wb.save(self._out_excel)
        self._wb.close()
        return


# Check File Encoding
def check_file_encoding(full_path_file):
    detector = UniversalDetector()
    with open(full_path_file, mode='rb') as f:
        for binary in f:
            detector.feed(binary)
            if detector.done:
                break
    detector.close()
    return detector.result, detector.result['encoding']


# Seek Directories
def seek_directories(write_excel: WriteExcel, level: int, dir_root: str, dir_relative: str, fp) -> None:

    dirs = []
    files = []

    for path in os.listdir(dir_root):
        if os.path.isfile(os.path.join(dir_root, path)):
            files.append(path)
        else:
            dirs.append(path)

    files.sort(key=str.lower)
    for file in files:
        base, ext = os.path.splitext(file)
        write_excel.write_cell(CELL_COL_NO, write_excel.get_count(), None, None, NUMBER_FORMAT)
        write_excel.write_cell(CELL_COL_PATH, dir_relative, ALIGN_LEFT_NO_WRAP, None, None)
        write_excel.write_cell(CELL_COL_FILE, file, ALIGN_LEFT_NO_WRAP, None, None)
        write_excel.write_cell(CELL_COL_EXT, ext, ALIGN_CENTER, None, None)
        # Ignore Files
        if (base.startswith('.') and ext == '') or ext in IGNORE_EXTENDS:
            atters = None
            encoding = None
        # Other Files
        else:
            atters, encoding = check_file_encoding(os.path.join(dir_root, file))
        if encoding is None:
            write_excel.write_cell(CELL_COL_ENCODING, '-', ALIGN_CENTER, None, None)
        else:
            write_excel.write_cell(CELL_COL_ENCODING, '%s' % encoding, ALIGN_LEFT_NO_WRAP, None, None)
        if atters is None:
            write_excel.write_cell(CELL_COL_ATTERS, '-', ALIGN_CENTER, None, None)
        else:
            write_excel.write_cell(CELL_COL_ATTERS, '%s' % atters, ALIGN_LEFT_NO_WRAP, None, None)
        print('%5d %s %s %s %s %s' %
              (write_excel.get_count(), dir_relative, file, ext,
               encoding if encoding is not None else '-', atters if atters is not None else '-'))
        fp.write('%5d %s %s %s %s %s\n' %
                 (write_excel.get_count(), dir_root, file, ext,
                  encoding if encoding is not None else '-', atters if atters is not None else '-'))

        write_excel.next_row()

    dirs.sort(key=str.lower)
    for dir_nest in dirs:
        seek_directories(write_excel, level + 1,
                         os.path.join(dir_root, dir_nest), os.path.join(dir_relative, dir_nest), fp)

    return


# Get Current Time
def get_current_time() -> str:

    now = datetime.datetime.now()
    dt = now.strftime("%Y-%m-%d %H:%M:%S")
    return dt


# Main
def main() -> None:

    try:
        options, arguments = getopt.getopt(sys.argv[1:], shortopts="h", longopts=["help"])
    except getopt.error as message:
        print(message)
        print(__doc__)
        sys.exit(1)

    for option, argument in options:
        if option in ("-h", "--help"):
            print(__doc__)
            sys.exit(0)

    print('File Encoding Checker - start [%s]' % get_current_time())

    # fp = None
    fp = open(OUT_DEBUG, 'w', encoding='utf-8')
    write_excel = WriteExcel(IN_EXCEL, OUT_EXCEL, OUT_SHEET)

    seek_directories(write_excel, 0, IN_SRC_ROOT + IN_SRC_RELATIVE, IN_SRC_RELATIVE, fp)

    write_excel.close()
    if fp is not None:
        fp.close()

    print('File Encoding Checker - end [%s]' % get_current_time())

    sys.exit(0)


# Goto Main
if __name__ == '__main__':
    main()
